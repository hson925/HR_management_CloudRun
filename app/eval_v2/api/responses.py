import logging
import uuid
import hashlib
import json as _json
import datetime as _dt
from concurrent.futures import ThreadPoolExecutor
from flask import request, session

logger = logging.getLogger(__name__)
from app.eval_v2.blueprints import eval_v2_api
from app.auth_utils import api_admin_required, api_role_required
from app.extensions import limiter
from app.eval_v2.api.common import (
    kst_now, _batch_delete, _VALID_EVAL_TYPES, _MAX_TEXT_LEN, _MAX_NAME_LEN,
    get_questions, get_weights,
    load_snapshot_questions, load_snapshot_weights,
    extract_valid_qids, extract_max_scores,
    EMP_ID_RE,
)
from app.services.firebase_service import get_firestore_client
from app.services.roster_cache_service import get_roster
from app.services.user_service import get_user_by_emp_id, get_user_by_email
from app.eval_v2.questions import EVAL_TYPE_LABELS
from app.utils.response import success, error
from app.utils.rate_limit import admin_rate_key
from app.constants import (
    CAMPUS_KO_TO_CODE, COL_EVAL_V2_SESSIONS, COL_EVAL_V2_RESPONSES,
)


_translation_pool = ThreadPoolExecutor(max_workers=3)


def _calc_payload_hash(scores, comment_en, comment_ko, open_answers):
    normalized = {k: float(v) for k, v in (scores or {}).items()}
    return hashlib.sha256(_json.dumps({
        's': normalized,
        'ce': comment_en or '',
        'ck': comment_ko or '',
        'o': open_answers or {},
    }, sort_keys=True, separators=(',', ':'), ensure_ascii=False).encode('utf-8')).hexdigest()


def _bg_translate_response(doc_id: str, open_answers: dict):
    """백그라운드 스레드에서 서술형 답변을 한국어·영어 양방향 번역 후 Firestore에 저장합니다."""
    try:
        from app.services.openai_service import translate_open_answers
        bilingual = translate_open_answers(open_answers)
        # {q_id: {'ko': ..., 'en': ...}} → 언어별 dict로 분리
        ko_map = {k: v['ko'] for k, v in bilingual.items()}
        en_map = {k: v['en'] for k, v in bilingual.items()}
        get_firestore_client().collection(COL_EVAL_V2_RESPONSES).document(doc_id).update({
            'open_answers_ko': ko_map,
            'open_answers_en': en_map,
            'translation_status': 'done',
            'translation_updated_at': kst_now(),
        })
    except Exception:
        logger.exception(f'_bg_translate_response failed [{doc_id}]')
        try:
            get_firestore_client().collection(COL_EVAL_V2_RESPONSES).document(doc_id).update({
                'translation_status': 'failed',
            })
        except Exception:
            logger.debug('Failed to mark translation_status=failed for doc %s', doc_id)


@eval_v2_api.route('/submit-eval', methods=['POST'])
@limiter.limit("5 per minute")
def api_submit_eval():
    try:
        data       = request.get_json(silent=True) or {}
        emp_id     = str(data.get('empId', '')).strip().lower()
        eval_type  = str(data.get('evalType', '')).strip().lower()
        rater_name = str(data.get('raterName', '')).strip()
        rater_role = str(data.get('raterRole', '')).strip()
        # rater_emp_id: 책임 추적 + dedup 정확도 향상.
        # 보안: portal 로그인 사용자는 session.emp_id 로 서버 강제 override (클라 위조 차단).
        # public_form / 외부 평가자 (비로그인) 는 빈문자 → fallback (rater_name 매칭) 그대로.
        sess_emp = str(session.get('emp_id', '')).strip().lower()
        if sess_emp:
            # admin 수동 입력 (is_manual) 일 때 admin 자신의 emp_id 가 들어가는 건 의도적 ×.
            # admin 은 다른 평가자를 대신 입력하므로 rater_emp_id 는 비워둠.
            if bool(data.get('isManual', False)):
                rater_emp_id = ''
            else:
                rater_emp_id = sess_emp
        else:
            rater_emp_id = ''
        # 안전 가드 — sess_emp 가 어떤 이유로 비정상이면 정규식 차단
        if rater_emp_id and not EMP_ID_RE.match(rater_emp_id):
            rater_emp_id = ''
        scores     = data.get('scores', {})
        comment_en   = str(data.get('commentEn', '')).strip()
        comment_ko   = str(data.get('commentKo', '')).strip()
        open_answers = data.get('openAnswers', {})

        if not emp_id or not eval_type or not rater_name or not rater_role:
            return error('Required fields missing.', 400)
        if eval_type not in _VALID_EVAL_TYPES:
            return error('Invalid eval type.', 400)
        if len(rater_name) > _MAX_NAME_LEN or len(rater_role) > _MAX_NAME_LEN:
            return error('Invalid field length.', 400)
        if len(comment_en) > _MAX_TEXT_LEN or len(comment_ko) > _MAX_TEXT_LEN:
            return error('Comment too long.', 400)
        if not isinstance(open_answers, dict):
            return error('Invalid open answers format.', 400)
        for ans_id, ans_text in open_answers.items():
            if len(str(ans_text)) > _MAX_TEXT_LEN:
                return error(f'Open answer too long: {ans_id}', 400)
        if not isinstance(scores, dict):
            return error('Invalid scores format.', 400)
        # 유효한 문항 ID 목록 조회 (fail-closed: 조회 실패 시 제출 차단)
        # 세션 스냅샷이 있으면 그것을 우선 사용 — api_get_questions_config와 동일한 로직
        try:
            session_id_for_qids = str(data.get('sessionId', '')).strip()
            snapshot = {}
            if session_id_for_qids:
                try:
                    db_qid = get_firestore_client()
                    sess_snap = db_qid.collection(COL_EVAL_V2_SESSIONS).document(session_id_for_qids).get()
                    if sess_snap.exists:
                        snapshot = sess_snap.to_dict().get('questions_snapshot', {})
                except Exception:
                    logger.debug('Session snapshot load failed for %s, falling back to global config', session_id_for_qids)
            roles_list = load_snapshot_questions(snapshot, eval_type)
            valid_qids = extract_valid_qids(roles_list)
            max_scores = extract_max_scores(roles_list)
            if not valid_qids:
                raise ValueError('Empty question list')
        except Exception:
            logger.exception(f'api_submit_eval: question config load failed ({eval_type})')
            return error('Could not load evaluation questions. Please try again.', 500)
        for qid, val in scores.items():
            if qid not in valid_qids:
                return error(f'Invalid question ID: {qid}', 400)
            try:
                fval = float(val)
                cap = max_scores.get(qid) or 5
                if fval != 0 and not (1 <= fval <= cap):
                    return error(f'Score out of range (1-{cap}): {qid}', 400)
            except (ValueError, TypeError):
                return error(f'Invalid score value: {qid}', 400)
        # openAnswers 키도 화이트리스트 검증
        for ans_id in open_answers:
            if ans_id not in valid_qids:
                return error(f'Invalid open answer ID: {ans_id}', 400)
        # 필수 서술형 문항 검증 — 비어있으면 차단
        # rater_role 에 해당하는 role 의 open_questions 중 required=True 인 것은 모두 채워져야 함
        required_oq_ids = []
        for role_obj in (roles_list or []):
            if not isinstance(role_obj, dict):
                continue
            role_name = role_obj.get('name') or role_obj.get('role', '')
            if role_name != rater_role:
                continue
            for oq in role_obj.get('open_questions', []) or []:
                if isinstance(oq, dict) and oq.get('required') and oq.get('id'):
                    required_oq_ids.append(oq['id'])
        for oq_id in required_oq_ids:
            if not str(open_answers.get(oq_id, '')).strip():
                return error(f'Required open answer missing: {oq_id}', 400)
        session_id = str(data.get('sessionId', '')).strip()
        if not session_id:
            return error('sessionId is required.', 400)
        db = get_firestore_client()
        is_manual = bool(data.get('isManual', False))
        # 수동 입력은 admin 전용 — 비인증 사용자가 세션 검증 우회 방지
        if is_manual and not session.get('admin_auth'):
            return error('Manual entry requires admin authentication.', 403)
        # 세션 존재 여부는 항상 검증
        sess_doc = db.collection(COL_EVAL_V2_SESSIONS).document(session_id).get()
        if not sess_doc.exists:
            return error('Session not found.', 404)
        # Passcode 게이트 (비-수동 입력만). admin 수동 입력은 admin_auth 가 보장되므
        # 로 내부 직원 면제가 자동 적용 — 이중 검증 무의미.
        if not is_manual:
            from app.eval_v2.api.passcode_gate import gate_check
            sess_snap_dict = dict(sess_doc.to_dict() or {})
            sess_snap_dict['id'] = session_id
            if not gate_check(sess_snap_dict):
                return error(
                    'Passcode required for this session.',
                    401,
                    code='PASSCODE_REQUIRED',
                )
        # 수동 입력이 아닌 경우에만 세션 상태 및 날짜 범위 검증
        # (admin 수동 입력은 세션 종료 후에도 허용)
        if not is_manual:
            if sess_doc.to_dict().get('status') != 'active':
                return error('Selected session is not active.', 400)
            # 세션 날짜 범위 체크 (KST 기준)
            sess_data = sess_doc.to_dict()
            start_date_str = sess_data.get('start_date', '')
            end_date_str = sess_data.get('end_date', '')
            try:
                _KST = _dt.timezone(_dt.timedelta(hours=9))
                today_kst = _dt.datetime.now(_KST).date()
                if start_date_str:
                    start_d = _dt.date.fromisoformat(start_date_str)
                    if today_kst < start_d:
                        return error(f'This session has not started yet. Opens {start_date_str}.', 400)
                if end_date_str:
                    end_d = _dt.date.fromisoformat(end_date_str)
                    if today_kst > end_d:
                        return error(f'This session period has ended. Closed {end_date_str}.', 400)
            except (ValueError, Exception):
                pass  # 날짜 파싱 실패 시 제한 없이 허용
        # 테스트 자동 태깅: emp_id가 test/demo/tmp 패턴이면 자동으로 is_test=True
        TEST_PATTERNS = ('test', 'demo', 'tmp', 'dummy')
        auto_test = any(emp_id.startswith(p) or p in emp_id for p in TEST_PATTERNS)
        # Doc ID 정책: 모든 제출 UUID. 슬롯 기반 dedup 폐기 — 다인 평가 (KT min_count=3)
        # 자유 누적을 위해. 같은 평가자의 재제출(잘못 입력 후 점수 수정 등) 점수 왜곡 방지는
        # 평균 산출 단계의 select_effective_responses() 가 같은 (역할, 정규화 rater_name)
        # 그룹의 submitted_at 최신 1건만 채택하는 방식으로 처리. payload_hash 필드는
        # forensic 용도로 그대로 저장 (비교에는 미사용).
        doc_id = str(uuid.uuid4())
        has_open = bool(open_answers and any(str(v).strip() for v in open_answers.values()))
        submitter = session.get('admin_email', 'public')
        # 제출자 IP 기록 — XFF-aware client_ip_key 공용 사용.
        from app.utils.rate_limit import client_ip_key
        submitter_ip = client_ip_key()
        # payload_hash: forensic / 디버그 용도. dedup 비교에 더 이상 사용 안 함.
        payload_hash = _calc_payload_hash(scores, comment_en, comment_ko, open_answers)
        payload = {
            'emp_id': emp_id, 'eval_type': eval_type,
            'rater_name': rater_name, 'rater_role': rater_role,
            'rater_emp_id': rater_emp_id,  # 빈문자 허용 (점진 누적 — public_form/외부평가자 fallback)
            'scores': scores, 'comment_en': comment_en, 'comment_ko': comment_ko,
            'open_answers': open_answers,
            'session_id': session_id,
            'is_test': auto_test,
            'is_manual': is_manual,
            'manual_reason': str(data.get('manualReason', '')).strip(),
            'submitted_at': kst_now(),
            'submitted_by': submitter,
            'submitted_ip': submitter_ip,
            'translation_status': 'pending' if has_open else 'skipped',
            'version': 1,
            'payload_hash': payload_hash,
        }
        ref = db.collection(COL_EVAL_V2_RESPONSES).document(doc_id)
        ref.set(payload)
        is_new_submission = True
        logger.info('eval submitted: emp=%s session=%s role=%s rater=%s by=%s ip=%s manual=%s',
                     emp_id, session_id, rater_role, rater_name, submitter, submitter_ip, is_manual)
        # 서술형 답변이 있으면 백그라운드에서 번역
        if has_open and is_new_submission:
            _translation_pool.submit(_bg_translate_response, doc_id, open_answers)
        return success({'docId': doc_id})
    except (ValueError, TypeError, KeyError) as e:
        # 입력 파싱/검증 실패 — 명확한 400.
        logger.warning('api_submit_eval input error: %s', e)
        return error('Invalid input.', 400)
    except Exception:
        # 예기치 않은 코드 버그 / Firestore 장애.
        logger.exception('api_submit_eval unexpected error')
        return error('An internal error occurred.', 500)


def _build_status_result(eval_type_filter, session_id_filter):
    """평가 제출 현황 집계 — 단일 코드 경로.
    반환: {campus_key: [{'id', 'name', 'campus', 'type', 'typeLabel', 'status', 'allDone'}]}
    get-status / get-campus-status 두 엔드포인트가 공유해 데이터 일관성 보장.
    """
    from collections import defaultdict
    db    = get_firestore_client()
    query = db.collection(COL_EVAL_V2_RESPONSES)
    if eval_type_filter:
        query = query.where('eval_type', '==', eval_type_filter)
    if session_id_filter:
        query = query.where('session_id', '==', session_id_filter)
    counts = defaultdict(lambda: defaultdict(int))
    raters = defaultdict(lambda: defaultdict(list))  # emp_id → role → [rater_name, ...]
    # 같은 (emp/역할/이름) 그룹의 최신 1건만 카운트 — 같은 평가자 재제출이 min_count 부풀림 방지.
    raw_docs = [doc.to_dict() for doc in query.limit(10000).stream()]
    from app.services.report_service import select_effective_responses
    for d in select_effective_responses(raw_docs):
        eid  = d.get('emp_id', '')
        role = d.get('rater_role', '')
        if not eid or not role:
            continue
        counts[eid][role] += 1
        name = str(d.get('rater_name', '')).strip()
        if name:
            raters[eid][role].append(name)

    snapshot = {}
    if session_id_filter:
        try:
            sess_doc = db.collection(COL_EVAL_V2_SESSIONS).document(session_id_filter).get()
            if sess_doc.exists:
                snapshot = sess_doc.to_dict().get('questions_snapshot', {})
        except Exception:
            logger.exception(f'_build_status_result: snapshot load failed [{session_id_filter}]')

    result = {}
    questions_cache = {}
    for row in get_roster():
        if len(row) < 4:
            continue
        eid   = str(row[2]).strip().lower()
        etype = str(row[3]).strip().lower()
        if eid in ('사번', ''):
            continue
        if eval_type_filter and etype != eval_type_filter:
            continue
        campus     = str(row[4]).strip() if len(row) > 4 else ''
        campus_key = campus if campus else 'Unknown'
        if etype not in questions_cache:
            questions_cache[etype] = load_snapshot_questions(snapshot, etype)
        roles = questions_cache[etype]
        if not roles or not isinstance(roles, list):
            continue
        role_status = []
        all_done    = True
        for role_obj in roles:
            if not isinstance(role_obj, dict):
                continue
            role_name = role_obj.get('name') or role_obj.get('role', '')
            if not role_name:
                continue
            # 표시 label 우선순위: eval admin_config 의 label_ko/label > portal_roles.label > raw name.
            # 두 시스템의 role 이름이 매칭될 때만 portal_roles fallback 적용. 매칭 실패 시 raw.
            from app.services import role_service as _rs
            role_label = (role_obj.get('label_ko')
                          or role_obj.get('label')
                          or _rs.get_role_label(role_name))
            min_count = role_obj.get('min_count', 1)
            current   = counts[eid].get(role_name, 0)
            done      = current >= min_count
            if not done:
                all_done = False
            role_status.append({
                'role': role_name, 'label': role_label,
                'current': current, 'required': min_count, 'done': done,
                'raters': raters[eid].get(role_name, []),
            })
        result.setdefault(campus_key, []).append({
            'id': eid, 'name': row[1] if len(row) > 1 else '',
            'campus': campus, 'type': etype,
            'typeLabel': EVAL_TYPE_LABELS.get(etype, etype.upper()),
            'status': role_status, 'allDone': all_done,
        })
    return result


@eval_v2_api.route('/get-status', methods=['POST'])
@api_admin_required
@limiter.limit("120 per minute", key_func=admin_rate_key)
def api_get_status():
    try:
        _body = request.get_json(silent=True) or {}
        result = _build_status_result(
            eval_type_filter=_body.get('evalType'),
            session_id_filter=str(_body.get('sessionId', '')).strip(),
        )
        return success({'data': result})
    except Exception:
        logger.exception('api_get_status error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/get-campus-status', methods=['POST'])
@api_role_required('GS', 'TL')
def api_get_campus_status():
    """GS/TL 용: 본인 캠퍼스 소속 피평가자 제출 여부만 반환.
    점수·코멘트·서술형 원문은 일절 포함하지 않음."""
    try:
        _body = request.get_json(silent=True) or {}

        # campus 는 세션이 아닌 Firestore 재조회로 결정 — 세션 변조 방어
        # GS 등 사번 없는 계정은 email fallback 으로 조회
        emp_id = str(session.get('emp_id', '')).strip().lower()
        me = get_user_by_emp_id(emp_id) if emp_id else None
        if not me:
            email = str(session.get('admin_email', '')).strip().lower()
            if email:
                me = get_user_by_email(email)
        if not me:
            return error('User record not found.', 404)
        my_campus = (me.get('campus') or '').strip()
        if not my_campus:
            return error(
                'Campus not assigned. Please contact admin.',
                403,
                code='NO_CAMPUS',
            )

        result = _build_status_result(
            eval_type_filter=_body.get('evalType'),
            session_id_filter=str(_body.get('sessionId', '')).strip(),
        )
        teachers = result.get(my_campus, [])
        return success({
            'campus': my_campus,
            'teachers': teachers,
        })
    except Exception:
        logger.exception('api_get_campus_status error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/get-responses', methods=['POST'])
@api_admin_required
@limiter.limit("120 per minute", key_func=admin_rate_key)
def api_get_responses():
    try:
        _body = request.get_json(silent=True) or {}
        emp_id = str(_body.get('empId', '')).strip().lower()
        session_id_filter = str(_body.get('sessionId', '')).strip()
        if not emp_id:
            return error('empId is required.', 400)
        db = get_firestore_client()
        docs = db.collection(COL_EVAL_V2_RESPONSES).where('emp_id', '==', emp_id).stream()
        responses = []
        for doc in docs:
            d = doc.to_dict()
            sid = d.get('session_id', '')
            if session_id_filter and sid != session_id_filter:
                continue
            responses.append({
                'doc_id': doc.id,
                'rater_name': d.get('rater_name', ''),
                'rater_role': d.get('rater_role', ''),
                'scores': d.get('scores', {}),
                'comment_en': d.get('comment_en', ''),
                'comment_ko': d.get('comment_ko', ''),
                'open_answers': d.get('open_answers', {}),
                'open_answers_ko': d.get('open_answers_ko', {}),
                'open_answers_en': d.get('open_answers_en', {}),
                'translation_status': d.get('translation_status', 'skipped'),
                'submitted_at': d.get('submitted_at', ''),
                'session_id': sid,
                'is_test': d.get('is_test', False),
                'is_manual': d.get('is_manual', False),
                'manual_reason': d.get('manual_reason', ''),
                'promoted_from_self_submit': d.get('promoted_from_self_submit', False),
                'version': int(d.get('version') or 0),
            })
        # 역할 → 문항 텍스트 맵 생성 (회차 스냅샷 우선 사용)
        eval_type = _body.get('evalType', '')
        questions_map = {}
        open_questions_map = {}
        snapshot_resp = {}
        if eval_type:
            if session_id_filter:
                try:
                    sess_doc = db.collection(COL_EVAL_V2_SESSIONS).document(session_id_filter).get()
                    if sess_doc.exists:
                        snapshot_resp = sess_doc.to_dict().get('questions_snapshot', {})
                except Exception:
                    logger.exception(f'Failed to load questions snapshot for responses [{session_id_filter}]')
            roles = load_snapshot_questions(snapshot_resp, eval_type)
            for r in roles:
                if isinstance(r, dict):
                    name = r.get('name') or r.get('role', '')
                    items = r.get('questions', r.get('items', []))
                    questions_map[name] = [
                        {'ko': q.get('text_ko') or q.get('ko', ''),
                         'en': q.get('text_en') or q.get('en', '')}
                        for q in items
                    ]
                    open_items = r.get('open_questions', [])
                    open_questions_map[name] = [
                        {'id': oq.get('id', ''),
                         'ko': oq.get('text_ko', ''),
                         'en': oq.get('text_en', '')}
                        for oq in open_items
                    ]
        weights_raw = {}
        if eval_type:
            weights_raw = load_snapshot_weights(snapshot_resp, eval_type)
        # 항상 0~1 float로 정규화
        weights_normalized = {}
        for k, v in weights_raw.items():
            v = float(v)
            if v > 1:
                v = v / 100
            weights_normalized[k] = round(v, 4)
        return success({
            'responses': responses,
            'questions': questions_map,
            'open_questions': open_questions_map,
            'weights': weights_normalized,
        })
    except Exception:
        logger.exception('api_get_responses error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/get-all-submissions', methods=['POST'])
@api_admin_required
@limiter.limit("60 per minute", key_func=admin_rate_key)
def api_get_all_submissions():
    try:
        db = get_firestore_client()
        rows = get_roster()
        # emp_id → emp_name, campus 맵 (단일 루프로 처리)
        name_map = {}
        campus_map = {}
        for row in rows:
            if len(row) > 2:
                eid = str(row[2]).strip().lower()
                name_map[eid] = row[1] if len(row) > 1 else ''
                if len(row) > 4:
                    campus_map[eid] = row[4]

        _body = request.get_json(silent=True) or {}
        session_id_filter = str(_body.get('session_id', _body.get('sessionId', ''))).strip()
        query = db.collection(COL_EVAL_V2_RESPONSES)
        if session_id_filter:
            query = query.where('session_id', '==', session_id_filter)
        docs = query.stream()
        submissions = []
        for doc in docs:
            d = doc.to_dict()
            sid = d.get('session_id', '')
            eid = d.get('emp_id', '')
            campus_ko = campus_map.get(eid, '')
            campus_key = campus_ko if not campus_ko.startswith('SUB') else 'SUB'
            campus_en_map = CAMPUS_KO_TO_CODE
            submissions.append({
                'doc_id': doc.id,
                'rater_name': d.get('rater_name', ''),
                'rater_role': d.get('rater_role', ''),
                'emp_id': eid,
                'emp_name': name_map.get(eid, ''),
                'emp_campus': campus_ko,
                'emp_campus_en': campus_en_map.get(campus_key, campus_ko),
                'eval_type': d.get('eval_type', ''),
                'submitted_at': d.get('submitted_at', ''),
                'session_id': sid,
                'is_test': d.get('is_test', False),
            })
        submissions.sort(key=lambda x: x.get('submitted_at',''), reverse=True)
        return success({'submissions': submissions})
    except Exception:
        logger.exception('api_get_all_submissions error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/export-csv', methods=['POST'])
@api_admin_required
@limiter.limit("10 per minute", key_func=admin_rate_key)
def api_export_csv():
    try:
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        data = request.get_json(silent=True) or {}
        session_id_filter = str(data.get('sessionId', '')).strip()
        include_test = bool(data.get('includeTest', False))
        db = get_firestore_client()

        # 로스터 맵
        rows = get_roster()
        name_map, campus_map = {}, {}
        for row in rows:
            if len(row) > 2:
                eid = str(row[2]).strip().lower()
                name_map[eid] = row[1] if len(row) > 1 else ''
                if len(row) > 4:
                    campus_map[eid] = row[4]

        # 응답 데이터 전체 수집
        query = db.collection(COL_EVAL_V2_RESPONSES)
        if session_id_filter:
            query = query.where('session_id', '==', session_id_filter)
        filtered = [
            doc.to_dict() for doc in query.limit(10000).stream()
            if include_test or not doc.to_dict().get('is_test', False)
        ]

        # 세션 스냅샷 로드 (문항 정의용)
        snapshot_csv = {}
        if session_id_filter:
            try:
                sess_doc = db.collection(COL_EVAL_V2_SESSIONS).document(session_id_filter).get()
                if sess_doc.exists:
                    snapshot_csv = sess_doc.to_dict().get('questions_snapshot', {})
            except Exception:
                logger.exception(f'Failed to load session snapshot for submissions [{session_id_filter}]')

        def get_q_defs(etype, role):
            roles_list = load_snapshot_questions(snapshot_csv, etype)
            for role_obj in (roles_list or []):
                if role_obj.get('name') == role:
                    return role_obj.get('questions', [])
            return []

        # 직책별로 그룹 (입력 순서 유지)
        sheets: dict = {}
        for d in filtered:
            role = d.get('rater_role', 'Unknown')
            sheets.setdefault(role, []).append(d)

        # Excel 생성
        wb = Workbook()
        wb.remove(wb.active)  # 기본 빈 시트 제거

        BASE_HEADERS = [
            'Session', 'Emp ID', 'Teacher Name', 'Campus', 'Eval Type',
            'Rater Name', 'Rater Role', 'Avg Score', 'Comment (EN)', 'Open Answers',
            'Is Manual', 'Manual Reason', 'Is Test', 'Submitted At',
        ]
        BASE_COL_WIDTHS = [18, 10, 16, 10, 10, 16, 12, 10, 25, 35, 9, 18, 9, 20]

        HDR_FILL  = PatternFill('solid', fgColor='1F2937')
        HDR_FONT  = Font(color='F9FAFB', bold=True, size=10)
        Q_FILL    = PatternFill('solid', fgColor='1E3A8A')
        Q_FONT    = Font(color='F9FAFB', bold=True, size=10)
        ALT_FILL  = PatternFill('solid', fgColor='F8FAFC')

        for role, role_rows in sheets.items():
            # Excel 시트명: 31자 제한, 특수문자 제거
            sheet_name = role[:31].translate(str.maketrans('/\\?*[]:', '-------'))
            ws = wb.create_sheet(title=sheet_name)

            # 이 시트의 eval_type 목록 (등장 순서 유지)
            seen_etypes: dict = {}
            for d in role_rows:
                seen_etypes.setdefault(d.get('eval_type', ''), None)
            etypes = list(seen_etypes.keys())
            multi_etype = len(etypes) > 1

            # 문항 컬럼 구성: (etype, q_id, header_label)
            q_columns: list = []
            for etype in etypes:
                qs = get_q_defs(etype, role)
                if qs:
                    for q in qs:
                        qid  = q.get('id', '')
                        text = q.get('text_en', qid)
                        label = (f"[{etype}] {qid}: {text}" if multi_etype
                                 else f"{qid}: {text}")
                        q_columns.append((etype, qid, label))
                else:
                    # 스냅샷 없을 때 실제 데이터에서 q_id 수집
                    for qid in sorted({
                        k for d in role_rows
                        if d.get('eval_type') == etype
                        for k in d.get('scores', {})
                    }):
                        label = (f"[{etype}] {qid}" if multi_etype else qid)
                        q_columns.append((etype, qid, label))

            # ── 헤더 행 ──
            all_headers = BASE_HEADERS + [col[2] for col in q_columns]
            ws.append(all_headers)
            ws.row_dimensions[1].height = 45
            for ci, _ in enumerate(all_headers, 1):
                cell = ws.cell(row=1, column=ci)
                cell.fill = HDR_FILL if ci <= len(BASE_HEADERS) else Q_FILL
                cell.font = HDR_FONT if ci <= len(BASE_HEADERS) else Q_FONT
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            # ── 데이터 행 ──
            for row_idx, d in enumerate(role_rows, 2):
                eid    = d.get('emp_id', '')
                scores = d.get('scores', {})
                vals   = [float(v) for v in scores.values() if v and float(v) > 0]
                avg    = round(sum(vals) / len(vals), 2) if vals else ''
                open_ans = '; '.join(
                    f"{k}: {v}" for k, v in (d.get('open_answers') or {}).items() if v
                )
                etype = d.get('eval_type', '')
                q_vals = [
                    scores.get(qid, '') if col_etype == etype else ''
                    for (col_etype, qid, _) in q_columns
                ]
                row_data = [
                    d.get('session_id', ''), eid.upper(),
                    name_map.get(eid, ''), campus_map.get(eid, ''),
                    etype, d.get('rater_name', ''), role, avg,
                    d.get('comment_en', ''), open_ans,
                    'Y' if d.get('is_manual') else 'N',
                    d.get('manual_reason', ''),
                    'Y' if d.get('is_test') else 'N',
                    d.get('submitted_at', ''),
                ] + q_vals
                ws.append(row_data)
                if row_idx % 2 == 0:
                    for ci in range(1, len(row_data) + 1):
                        ws.cell(row=row_idx, column=ci).fill = ALT_FILL

            # ── 컬럼 너비 ──
            for ci, w in enumerate(BASE_COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            for ci in range(len(BASE_HEADERS) + 1, len(all_headers) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 40

            # ── 틀 고정 (헤더 행) ──
            ws.freeze_panes = 'A2'

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import Response
        import re as _re
        safe_sid = _re.sub(r'[^a-zA-Z0-9_\-]', '', session_id_filter) if session_id_filter else 'all'
        filename = f"eval_v2_{safe_sid}_{kst_now()[:10]}.xlsx"
        return Response(
            buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception:
        logger.exception('api_export_csv error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/update-eval', methods=['POST'])
@api_admin_required
@limiter.limit("30 per minute", key_func=admin_rate_key)
def api_update_eval():
    try:
        data     = request.get_json(silent=True) or {}
        doc_id   = str(data.get('docId', '')).strip()
        scores   = data.get('scores', {})
        comment_en = str(data.get('commentEn', '')).strip()
        comment_ko = str(data.get('commentKo', '')).strip()
        rater_name = str(data.get('raterName', '')).strip()
        reason     = str(data.get('manualReason', '')).strip()
        open_answers = data.get('openAnswers', {})
        if not doc_id:
            return error('docId is required.', 400)
        # H2: rater_name 길이 검증 (submit과 동일 기준)
        if len(rater_name) > _MAX_NAME_LEN:
            return error('Invalid field length.', 400)
        # H3: comment 길이 검증 (submit과 동일 기준)
        if len(comment_en) > _MAX_TEXT_LEN or len(comment_ko) > _MAX_TEXT_LEN:
            return error('Comment too long.', 400)
        # open_answers 검증
        if not isinstance(open_answers, dict):
            return error('Invalid open answers format.', 400)
        for ans_id, ans_text in open_answers.items():
            if len(str(ans_text)) > _MAX_TEXT_LEN:
                return error(f'Open answer too long: {ans_id}', 400)
        # H1: 점수 검증 — submit과 동일하게 0(미응답) 허용, 그 외 1..max_score(per-qid)
        if not isinstance(scores, dict):
            return error('Invalid scores format.', 400)
        # per-qid max_score 로드: 트랜잭션 외부 read 1회 (response 문서 → session_id/eval_type → snapshot).
        # NOTE: 검증↔트랜잭션 사이 admin 이 save-session-questions 로 max_score 를 축소해도
        # 이 update 는 stale cap 으로 통과될 수 있음. 무해 — 데이터 손상 없음, 단지 일시적으로
        # 더 너그러운 검증. 다음 update 부터 새 cap 적용.
        max_scores = {}
        required_oq_ids = []
        try:
            db_pre = get_firestore_client()
            pre_snap = db_pre.collection(COL_EVAL_V2_RESPONSES).document(doc_id).get()
            if pre_snap.exists:
                pre_data = pre_snap.to_dict() or {}
                pre_session_id = pre_data.get('session_id')
                pre_eval_type = pre_data.get('eval_type')
                pre_rater_role = pre_data.get('rater_role')
                snapshot = {}
                if pre_session_id:
                    sess_snap = db_pre.collection(COL_EVAL_V2_SESSIONS).document(pre_session_id).get()
                    if sess_snap.exists:
                        snapshot = sess_snap.to_dict().get('questions_snapshot', {})
                if pre_eval_type:
                    pre_roles = load_snapshot_questions(snapshot, pre_eval_type)
                    max_scores = extract_max_scores(pre_roles)
                    # 동일 rater_role 의 required OQ id 추출 (submit 과 동일 규칙)
                    for role_obj in (pre_roles or []):
                        if not isinstance(role_obj, dict):
                            continue
                        role_name = role_obj.get('name') or role_obj.get('role', '')
                        if role_name != pre_rater_role:
                            continue
                        for oq in role_obj.get('open_questions', []) or []:
                            if isinstance(oq, dict) and oq.get('required') and oq.get('id'):
                                required_oq_ids.append(oq['id'])
        except Exception:
            logger.debug('api_update_eval: max_scores/required-oq load failed for %s, falling back', doc_id)
        for qid, val in scores.items():
            try:
                fval = float(val)
                cap = max_scores.get(qid) or 5
                if fval != 0 and not (1 <= fval <= cap):
                    return error(f'Scores must be 0 or in range 1-{cap}. ({qid}: {val})', 400)
            except (ValueError, TypeError):
                return error(f'Invalid score format. ({qid}: {val})', 400)
        # 필수 서술형 검증 (submit 과 동일 정책)
        for oq_id in required_oq_ids:
            if not str(open_answers.get(oq_id, '')).strip():
                return error(f'Required open answer missing: {oq_id}', 400)
        has_open = bool(open_answers and any(str(v).strip() for v in open_answers.values()))
        db = get_firestore_client()
        # Optimistic locking: 클라이언트가 보내온 version과 현재 저장된 version을
        # 트랜잭션 안에서 비교해 다른 admin의 변경 덮어쓰기를 방지.
        # client_version 미전송(legacy 클라이언트) 시 검사 skip.
        client_version_raw = data.get('version', None)
        try:
            client_version = int(client_version_raw) if client_version_raw is not None else None
        except (ValueError, TypeError):
            client_version = None
        # payload hash 재계산 (submit과 동일 규칙) — 수정본도 idempotent 비교용
        new_payload_hash = _calc_payload_hash(scores, comment_en, comment_ko, open_answers)
        base_update = {
            'scores': scores,
            'comment_en': comment_en,
            'comment_ko': comment_ko,
            'rater_name': rater_name,
            'manual_reason': reason,
            'open_answers': open_answers,
            'is_manual': True,
            'updated_at': kst_now(),
            'updated_by': session.get('admin_email', ''),
            'payload_hash': new_payload_hash,
        }
        if has_open:
            base_update['translation_status'] = 'pending'

        from google.cloud import firestore as _fs

        class _VersionConflict(Exception):
            def __init__(self, current_version):
                self.current_version = current_version

        class _NotFound(Exception):
            pass

        ref = db.collection(COL_EVAL_V2_RESPONSES).document(doc_id)

        @_fs.transactional
        def _update_txn(tx):
            snap = ref.get(transaction=tx)
            if not snap.exists:
                raise _NotFound()
            existing = snap.to_dict() or {}
            stored_version = int(existing.get('version') or 0)
            # client_version이 제공된 경우에만 충돌 검사
            if client_version is not None and client_version != stored_version:
                raise _VersionConflict(stored_version)
            new_version = stored_version + 1
            tx.update(ref, {**base_update, 'version': new_version})
            return new_version

        try:
            new_version = _update_txn(db.transaction())
        except _NotFound:
            return error('Response not found.', 404)
        except _VersionConflict as e:
            return error(
                'This evaluation was modified by another administrator. Please reload and try again.',
                409,
                code='VERSION_CONFLICT',
                currentVersion=e.current_version,
            )

        if has_open:
            _translation_pool.submit(_bg_translate_response, doc_id, open_answers)
        from app.services.audit_service import log_audit
        log_audit('eval_response_update', session.get('admin_email', ''), target=doc_id, details={'reason': reason, 'version': new_version}, category='response')
        return success({'version': new_version})
    except Exception:
        logger.exception('api_update_eval error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/delete-response', methods=['POST'])
@api_admin_required
@limiter.limit("30 per minute", key_func=admin_rate_key)
def api_delete_response():
    try:
        _body = request.get_json(silent=True) or {}
        doc_id = str(_body.get('docId', '')).strip()
        if not doc_id:
            return error('docId is required.', 400)
        db = get_firestore_client()
        db.collection(COL_EVAL_V2_RESPONSES).document(doc_id).delete()
        from app.services.audit_service import log_audit
        log_audit('eval_response_delete', session.get('admin_email', ''), target=doc_id, category='response')
        return success()
    except Exception:
        logger.exception('api_delete_response error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/mark-test', methods=['POST'])
@api_admin_required
@limiter.limit("60 per minute", key_func=admin_rate_key)
def api_mark_test():
    try:
        _body = request.get_json(silent=True) or {}
        doc_id  = str(_body.get('docId', '')).strip()
        is_test = bool(_body.get('isTest', True))
        if not doc_id:
            return error('docId is required.', 400)
        db = get_firestore_client()
        db.collection(COL_EVAL_V2_RESPONSES).document(doc_id).update({'is_test': is_test})
        return success()
    except Exception:
        logger.exception('api_mark_test error')
        return error('An internal error occurred.', 500)


class _PromoteNotFound(Exception):
    pass


class _PromoteAlreadyPromoted(Exception):
    pass


class _PromoteAlreadyManual(Exception):
    pass


class _PromoteIsTest(Exception):
    pass


class _DepromoteNotPromoted(Exception):
    pass


@eval_v2_api.route('/promote-response', methods=['POST'])
@api_admin_required
@limiter.limit("30 per minute", key_func=admin_rate_key)
def api_promote_response():
    """동명이인 등으로 평균 미채택된 self-submit 응답을 manual entry 로 승격.

    is_manual=True 로 플래그 변경 → select_effective_responses 의 manual 분기
    (doc_id 별 독립 키) 가 dedup 무관하게 채택. 점수/코멘트/서술형은 보존.
    promoted_* 메타로 추적·undo 가능.

    트랜잭션으로 가드/version 업데이트를 atomic 하게 처리 — 동시 update-eval /
    depromote / 다른 promote 와의 race 차단.
    """
    try:
        _body = request.get_json(silent=True) or {}
        doc_id = str(_body.get('docId', '')).strip()
        if not doc_id:
            return error('docId is required.', 400)
        db = get_firestore_client()
        ref = db.collection(COL_EVAL_V2_RESPONSES).document(doc_id)

        from google.cloud import firestore as _fs

        @_fs.transactional
        def _promote_txn(tx):
            snap = ref.get(transaction=tx)
            if not snap.exists:
                raise _PromoteNotFound()
            existing = snap.to_dict() or {}
            if existing.get('promoted_from_self_submit'):
                raise _PromoteAlreadyPromoted()
            if existing.get('is_manual'):
                raise _PromoteAlreadyManual()
            if existing.get('is_test'):
                raise _PromoteIsTest()
            stored_version = int(existing.get('version') or 0)
            tx.update(ref, {
                'is_manual': True,
                'manual_reason': existing.get('manual_reason')
                    or 'Promoted from self-submit due to dup-name resolution',
                'promoted_from_self_submit': True,
                'promoted_at': kst_now(),
                'promoted_by': session.get('admin_email', ''),
                'version': stored_version + 1,
            })

        try:
            _promote_txn(db.transaction())
        except _PromoteNotFound:
            return error('Response not found.', 404)
        except _PromoteAlreadyPromoted:
            return error('Already promoted.', 400)
        except _PromoteAlreadyManual:
            return error('Already a manual entry.', 400)
        except _PromoteIsTest:
            return error('Cannot promote a test response. Remove the test mark first.', 400)

        from app.services.audit_service import log_audit
        log_audit('eval_response_promote', session.get('admin_email', ''),
                  target=doc_id, category='response')
        return success()
    except (ValueError, TypeError, KeyError) as e:
        logger.warning('api_promote_response input error: %s', e)
        return error('Invalid input.', 400)
    except Exception:
        logger.exception('api_promote_response error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/depromote-response', methods=['POST'])
@api_admin_required
@limiter.limit("30 per minute", key_func=admin_rate_key)
def api_depromote_response():
    """promoted self-submit 응답을 원래 self-submit 상태로 되돌림.

    가드: promoted_from_self_submit=True 인 응답만 허용. 진짜 manual entry
    (admin 이 직접 입력한) 는 보호 — undo 시 self-submit 으로 변환되면 데이터
    의미가 왜곡됨.

    update-eval 로 admin 이 점수를 수정한 promoted 응답이라도 depromote 가능.
    수정된 점수는 그대로 보존 (is_manual 만 토글). 그 응답은 다시 rater_name
    그룹의 dedup 에 들어가므로, 같은 이름의 더 최신 응답이 있으면 ⛔ 처리.

    트랜잭션으로 가드/version 업데이트를 atomic 하게 처리 — 동시 update-eval /
    promote / 다른 depromote 와의 race 차단.
    """
    try:
        _body = request.get_json(silent=True) or {}
        doc_id = str(_body.get('docId', '')).strip()
        if not doc_id:
            return error('docId is required.', 400)
        db = get_firestore_client()
        ref = db.collection(COL_EVAL_V2_RESPONSES).document(doc_id)

        from google.cloud import firestore as _fs

        @_fs.transactional
        def _depromote_txn(tx):
            snap = ref.get(transaction=tx)
            if not snap.exists:
                raise _PromoteNotFound()
            existing = snap.to_dict() or {}
            if not existing.get('promoted_from_self_submit'):
                raise _DepromoteNotPromoted()
            stored_version = int(existing.get('version') or 0)
            tx.update(ref, {
                'is_manual': False,
                'promoted_from_self_submit': _fs.DELETE_FIELD,
                'promoted_at': _fs.DELETE_FIELD,
                'promoted_by': _fs.DELETE_FIELD,
                'manual_reason': _fs.DELETE_FIELD,
                'depromoted_at': kst_now(),
                'depromoted_by': session.get('admin_email', ''),
                'version': stored_version + 1,
            })

        try:
            _depromote_txn(db.transaction())
        except _PromoteNotFound:
            return error('Response not found.', 404)
        except _DepromoteNotPromoted:
            return error(
                'This response was not promoted from self-submit; depromote is blocked '
                'to avoid data semantic loss.', 400)

        from app.services.audit_service import log_audit
        log_audit('eval_response_depromote', session.get('admin_email', ''),
                  target=doc_id, category='response')
        return success()
    except (ValueError, TypeError, KeyError) as e:
        logger.warning('api_depromote_response input error: %s', e)
        return error('Invalid input.', 400)
    except Exception:
        logger.exception('api_depromote_response error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/delete-test-responses', methods=['POST'])
@api_admin_required
@limiter.limit("10 per minute", key_func=admin_rate_key)
def api_delete_test_responses():
    try:
        _body = request.get_json(silent=True) or {}
        session_id = str(_body.get('sessionId', '')).strip()
        db = get_firestore_client()
        query = db.collection(COL_EVAL_V2_RESPONSES).where('is_test', '==', True)
        if session_id:
            query = query.where('session_id', '==', session_id)
        docs = list(query.stream())
        count = _batch_delete(db, docs)
        from app.services.audit_service import log_audit
        log_audit('eval_test_responses_delete', session.get('admin_email', ''), target=session_id or 'all', details={'deleted_count': count}, category='response')
        return success({'deletedCount': count})
    except Exception:
        logger.exception('api_delete_test_responses error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/responses/<doc_id>/translate', methods=['POST'])
@api_admin_required
@limiter.limit("20 per minute", key_func=admin_rate_key)
def api_retranslate_response(doc_id):
    """서술형 답변 번역을 수동으로 재실행합니다 (번역 실패 시 재시도용)."""
    try:
        db = get_firestore_client()
        doc = db.collection(COL_EVAL_V2_RESPONSES).document(doc_id).get()
        if not doc.exists:
            return error('Response not found.', 404)
        open_answers = doc.to_dict().get('open_answers', {})
        if not open_answers or not any(str(v).strip() for v in open_answers.values()):
            return error('No open answers to translate.', 400)
        db.collection(COL_EVAL_V2_RESPONSES).document(doc_id).update({'translation_status': 'pending'})
        _translation_pool.submit(_bg_translate_response, doc_id, open_answers)
        return success({'message': 'Translation started in background.'})
    except Exception:
        logger.exception('api_retranslate_response error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/responses/<doc_id>/update-translation', methods=['POST'])
@api_admin_required
@limiter.limit("30 per minute", key_func=admin_rate_key)
def api_update_translation(doc_id):
    """관리자가 수동으로 수정한 번역 내용을 저장합니다."""
    try:
        data = request.get_json(silent=True) or {}
        updated_ko = data.get('open_answers_ko', {})
        updated_en = data.get('open_answers_en', {})
        if not isinstance(updated_ko, dict) or not isinstance(updated_en, dict):
            return error('open_answers_ko and open_answers_en must be dicts.', 400)
        update_payload = {
            'translation_status': 'done',
            'translation_updated_at': kst_now(),
            'translation_edited_by': session.get('admin_email', ''),
        }
        if updated_ko:
            update_payload['open_answers_ko'] = updated_ko
        if updated_en:
            update_payload['open_answers_en'] = updated_en
        db = get_firestore_client()
        db.collection(COL_EVAL_V2_RESPONSES).document(doc_id).update(update_payload)
        return success({'message': 'Translation saved.'})
    except Exception:
        logger.exception('api_update_translation error')
        return error('An internal error occurred.', 500)


@eval_v2_api.route('/pending-translations', methods=['POST'])
@api_admin_required
@limiter.limit("60 per minute", key_func=admin_rate_key)
def api_pending_translations():
    """번역 대기 중인 서술형 응답 목록.

    submitted_at 기준 5분 이상 pending 상태인 항목만 반환.
    관리자가 stale pending 응답을 확인하고 재시도할 수 있도록 제공.
    """
    try:
        cutoff = (
            _dt.datetime.now(_dt.timezone.utc) - _dt.timedelta(minutes=5)
        ).isoformat()
        db = get_firestore_client()
        docs = (db.collection(COL_EVAL_V2_RESPONSES)
                .where('translation_status', '==', 'pending')
                .stream())
        stale = []
        for doc in docs:
            d = doc.to_dict()
            submitted_at = d.get('submitted_at', '')
            if submitted_at and submitted_at < cutoff:
                stale.append({
                    'doc_id':       doc.id,
                    'emp_id':       d.get('emp_id', ''),
                    'rater_role':   d.get('rater_role', ''),
                    'submitted_at': submitted_at,
                    'session_id':   d.get('session_id', ''),
                })
        return success({'items': stale, 'count': len(stale)})
    except Exception:
        logger.exception('api_pending_translations error')
        return error('An internal error occurred.', 500)
